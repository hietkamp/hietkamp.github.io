#!/usr/bin/env python3
"""
EA Alpha Assessment — Nederlandse versie.
Structuur + opmaak identiek aan het origineel; alleen de 12 alphas van de huidige site.

Uitvoeren:  .venv/bin/python build/build_xlsx.py
Vereist:    openpyxl  (.venv/bin/pip install openpyxl)
Output:     docs/EA_Alpha_Assessment.xlsx

Wijzig dit bestand en draai opnieuw — bewerk de xlsx nooit met de hand.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Vaste kleuren ────────────────────────────────────────────────────────────
INK    = "1C1A16"
DARK   = "1C2433"   # dashboard-achtergrond / koptekst alpha-tab
PAPER  = "F3EFE6"
PAPER2 = "F4F2EC"   # zebra-rij (warm grijs) — exact origineel
WHITE  = "FFFFFF"
GRAY   = "5B6472"   # ondertitels / labels
CUST   = "2E7D52"
SOL    = "B07D12"
END    = "2A5E9C"
MVA_C  = "0F6E63"
MVG_C  = "4B3F9E"
KRN_C  = "27406B"
CHK_GR = "1E7A4D"   # ✔ groen (checkbox + bereikt status)
AMB_BG = "FBF0D8"   # amber achtergrond toestandsstatus
AMB_FG = "8A6D1A"   # amber tekst toestandsstatus

OWN_COL  = {"MVA": MVA_C, "MVG": MVG_C, "Kernel": KRN_C}
LVL_COL  = {"Enterprise": KRN_C, "Solution": MVA_C, "Spans": MVG_C}
LVL_LBL  = {"Enterprise": "ENTERPRISE", "Solution": "SOLUTION",
             "Spans": "SPANS LEVELS"}
AREA_COL = {"cust": CUST, "sol": SOL, "end": END}


def tint(hex6, t=0.857):
    """Meng hex_kleur met wit op factor t (0=origineel, 1=wit)."""
    r = int(hex6[0:2], 16)
    g = int(hex6[2:4], 16)
    b = int(hex6[4:6], 16)
    return (f"{round(r+(255-r)*t):02X}"
            f"{round(g+(255-g)*t):02X}"
            f"{round(b+(255-b)*t):02X}")


# ─── Alpha-definities ─────────────────────────────────────────────────────────
# states = [(staatsnaam_EN, [checklistpunten_NL]), ...]
# sample = (volle_toestanden, extra_punten_in_volgende)
ALPHAS = [
    # ── CUSTOMER ──────────────────────────────────────────────────────────
    dict(name="Opportunity", tab="Opportunity",
         owner="Kernel", area="cust", level="Solution",
         target="Benefit Accrued",
         desc="De omstandigheden die een nieuw of gewijzigd systeem de moeite waard maken.",
         states=[
             ("Identified", [
                 "Een idee voor het verbeteren van de manier van werken, het vergroten van het marktaandeel of het toepassen van een nieuw systeem is geïdentificeerd.",
                 "Minimaal één stakeholder wil investeren in het begrijpen van de opportunity en de waarde ervan.",
                 "De overige stakeholders die de opportunity delen zijn geïdentificeerd.",
             ]),
             ("Solution Needed", [
                 "De stakeholders bij de opportunity en de voorgestelde oplossing zijn geïdentificeerd.",
                 "De behoeften van stakeholders die de opportunity genereren zijn vastgesteld.",
                 "Eventuele onderliggende problemen en hun oorzaken zijn geïdentificeerd.",
                 "Bevestigd is dat een systeem-gebaseerde oplossing nodig is.",
                 "Minimaal één oplossing is voorgesteld.",
             ]),
             ("Value Established", [
                 "De waarde van het aanpakken van de opportunity is gekwantificeerd.",
                 "De impact van de oplossing op de stakeholders is begrepen.",
                 "De waarde die het systeem biedt aan degenen die het financieren en gebruiken is begrepen.",
                 "De succescriteria voor de inzet zijn duidelijk.",
                 "De gewenste uitkomsten zijn duidelijk en gekwantificeerd.",
             ]),
             ("Viable", [
                 "Een oplossing is geschetst.",
                 "De oplossing kan worden ontwikkeld en ingezet binnen de beperkingen.",
                 "De risico's van de oplossing zijn aanvaardbaar en beheersbaar.",
                 "De globale kosten zijn lager dan de verwachte waarde van de opportunity.",
                 "Het is duidelijk dat het nastreven van de opportunity haalbaar is.",
             ]),
             ("Addressed", [
                 "Een bruikbaar systeem dat aantoonbaar de opportunity aanpakt is beschikbaar.",
                 "Stakeholders zijn het erover eens dat de beschikbare oplossing het waard is om in te zetten.",
                 "Stakeholders zijn tevreden dat de oplossing de opportunity aanpakt.",
             ]),
             ("Benefit Accrued", [
                 "De oplossing heeft voordelen voor stakeholders beginnen op te leveren.",
                 "Het rendement op investering is minimaal zo goed als verwacht.",
             ]),
         ], sample=(3, 2)),

    dict(name="Stakeholders", tab="Stakeholders",
         owner="Kernel", area="cust", level="Solution",
         target="Satisfied in Use",
         desc="De mensen en organisaties die het systeem beïnvloeden of erdoor worden beïnvloed.",
         states=[
             ("Recognized", [
                 "De stakeholdergroepen zijn geïdentificeerd.",
                 "De sleutelstakeholdergroepen zijn vertegenwoordigd.",
                 "De verantwoordelijkheden van de stakeholdervertegenwoordigers zijn gedefinieerd.",
             ]),
             ("Represented", [
                 "De vertegenwoordigers hebben hun verantwoordelijkheden overeengekomen.",
                 "De vertegenwoordigers zijn gemachtigd hun verantwoordelijkheden uit te voeren.",
                 "De samenwerkingsaanpak onder vertegenwoordigers is overeengekomen.",
                 "De manier van werken van het team wordt ondersteund en gerespecteerd door de stakeholders.",
             ]),
             ("Involved", [
                 "De vertegenwoordigers ondersteunen het team conform hun verantwoordelijkheden.",
                 "De vertegenwoordigers geven feedback en nemen tijdig beslissingen.",
                 "Wijzigingen worden tijdig gecommuniceerd aan het team en stakeholders.",
             ]),
             ("In Agreement", [
                 "De minimale verwachtingen van de stakeholdergroepen zijn overeengekomen.",
                 "De vertegenwoordigers zijn tevreden met hun betrokkenheid.",
                 "De bijdrage van de vertegenwoordigers wordt gewaardeerd door het team.",
                 "De bijdrage van het team wordt gewaardeerd door de vertegenwoordigers.",
                 "Prioriteiten zijn duidelijk en perspectieven zijn gebalanceerd.",
             ]),
             ("Satisfied for Deployment", [
                 "De vertegenwoordigers geven feedback op het systeem vanuit het perspectief van hun groep.",
                 "De vertegenwoordigers bevestigen dat het systeem klaar is voor inzet.",
             ]),
             ("Satisfied in Use", [
                 "De stakeholders gebruiken het systeem en geven feedback.",
                 "De stakeholders bevestigen dat het systeem aan hun verwachtingen voldoet.",
             ]),
         ], sample=(3, 3)),

    # ── SOLUTION ──────────────────────────────────────────────────────────
    dict(name="Requirements", tab="Requirements",
         owner="Kernel", area="sol", level="Solution",
         target="Fulfilled",
         desc="Wat het systeem moet doen om de opportunity aan te pakken en de stakeholders tevreden te stellen.",
         states=[
             ("Conceived", [
                 "De initiële stakeholders zijn het erover eens dat een systeem geproduceerd moet worden.",
                 "De stakeholders die het systeem zullen gebruiken zijn geïdentificeerd.",
                 "De stakeholders die het initiële werk zullen financieren zijn geïdentificeerd.",
                 "Er is een duidelijke opportunity voor het systeem om aan te pakken.",
             ]),
             ("Bounded", [
                 "De stakeholders betrokken bij de ontwikkeling van het systeem zijn geïdentificeerd.",
                 "De stakeholders zijn het eens over het doel van het systeem.",
                 "Het is duidelijk wat succes is voor het systeem.",
                 "Er is een gedeeld begrip van de omvang van de oplossing.",
                 "De manier waarop requirements beschreven worden is overeengekomen.",
                 "De mechanismen voor het beheer van requirements zijn aanwezig.",
                 "Het prioriteringsschema is duidelijk.",
                 "Beperkingen zijn geïdentificeerd en meegewogen.",
                 "Aannames zijn duidelijk geformuleerd.",
             ]),
             ("Coherent", [
                 "De requirements zijn vastgelegd en gedeeld met het team en stakeholders.",
                 "De herkomst en onderbouwing van de requirements zijn duidelijk.",
                 "Conflicterende requirements zijn geïdentificeerd en aangepakt.",
                 "De requirements beschrijven de essentiële kenmerken van het systeem.",
                 "De belangrijkste gebruiksscenario's kunnen worden uitgelegd.",
                 "De prioriteit van de requirements is duidelijk.",
                 "Het team begrijpt wat opgeleverd moet worden en is het erover eens dit te doen.",
             ]),
             ("Acceptable", [
                 "Stakeholders accepteren dat de requirements een aanvaardbare oplossing beschrijven.",
                 "Het wijzigingstempo van overeengekomen requirements is laag en onder controle.",
                 "De waarde van het implementeren van de requirements is duidelijk.",
                 "De requirements zijn testbaar.",
             ]),
             ("Addressed", [
                 "Genoeg requirements zijn afgehandeld voor het systeem om aanvaardbaar te zijn.",
                 "De geïmplementeerde requirements leveren duidelijke waarde.",
                 "Het systeem wordt door stakeholders geaccepteerd als het waard om operationeel te maken.",
             ]),
             ("Fulfilled", [
                 "Stakeholders accepteren dat de requirements vastleggen wat zij nodig hebben.",
                 "Geen openstaande requirements verhinderen acceptatie.",
                 "Het systeem voldoet volledig aan de requirements.",
             ]),
         ], sample=(2, 4)),

    dict(name="System", tab="System",
         owner="Kernel", area="sol", level="Solution",
         target="Operational",
         desc="Het systeem dat waarde levert — software, geconfigureerde producten, hardware, data en services. Een generalisatie van Essence's Software System.",
         states=[
             ("Architecture Selected", [
                 "De criteria voor het selecteren van de architectuur zijn overeengekomen.",
                 "Een architectuur is geselecteerd die de sleuteltechnische risico's aanpakt.",
                 "De bouwen / kopen / configureren-beslissingen zijn genomen.",
                 "De technologie- en platformbeslissingen zijn overeengekomen met stakeholders.",
             ]),
             ("Demonstrable", [
                 "De sleutelarchitectuurkenmerken zijn aangetoond.",
                 "De relevante stakeholders zijn het erover eens dat de architectuur geschikt is.",
                 "De kritieke interfaces en integraties zijn getest.",
                 "Prestaties zijn gemeten ten opzichte van de sleuteldrivers.",
             ]),
             ("Usable", [
                 "Het systeem kan worden bediend door representatieve gebruikers.",
                 "De functionaliteit is aangetoond.",
                 "De prestatie- en defectniveaus zijn aanvaardbaar.",
                 "Een bruikbaar systeem is beschikbaar.",
             ]),
             ("Ready", [
                 "Het systeem heeft de overeengekomen tests doorstaan.",
                 "Operationele en ondersteuningsdocumentatie is beschikbaar.",
                 "De stakeholdervertegenwoordigers accepteren het systeem.",
                 "Het systeem is klaar voor inzet.",
             ]),
             ("Operational", [
                 "Het systeem is in gebruik in de operationele omgeving.",
                 "Het systeem wordt ondersteund conform de overeengekomen serviceniveaus.",
                 "Het systeem is beschikbaar voor de beoogde gebruikers.",
             ]),
             ("Retired", [
                 "Het systeem wordt niet meer ondersteund.",
                 "Het systeem is vervangen of uit gebruik genomen.",
             ]),
         ], sample=(1, 2)),

    dict(name="Architectural Drivers", tab="Architectural Drivers",
         owner="MVA", area="sol", level="Solution",
         target="Sustained",
         desc=("De zakelijke en technische eisen die architectuurkeuzes sturen. "
               "Omvat: strategische drivers (aanleiding), doelen (goals), "
               "kwaliteitsattributen (performance, security, …), beperkingen (constraints) "
               "en principes. Vergelijkbaar met de ArchiMate motivatielaag "
               "(Driver · Goal · Requirement · Constraint · Principle)."),
         states=[
             ("Identified", [
                 "Architectureel significante drivers zijn naar voren gebracht — "
                 "denk aan strategische doelen, kwaliteitsattributen (bijv. performance, "
                 "beveiliging), beperkingen (budget, technologie) en geldende principes.",
                 "Van elke driver zijn type, omschrijving en eigenaar vastgelegd.",
             ]),
             ("Prioritised", [
                 "De architectureel significante drivers zijn overeengekomen en "
                 "naar gewicht gerangschikt.",
                 "De overige drivers zijn expliciet uitgesteld of verworpen.",
             ]),
             ("Quantified", [
                 "Elke significante driver heeft een meetbaar succescriterium "
                 "(bijv. responstijd < 500 ms, beschikbaarheid ≥ 99,9%, kosten < X).",
                 "De meetcriteria zijn overeengekomen met de stakeholders.",
             ]),
             ("Addressed", [
                 "De drivers zijn weerspiegeld in de beslissingen en de architectuur.",
                 "Bewijs toont aan dat de drivers worden nageleefd.",
             ]),
             ("Sustained", [
                 "De drivers worden bewaakt naarmate de context verandert.",
                 "De drivers worden opnieuw gevalideerd en bijgewerkt wanneer nodig.",
             ]),
         ], sample=(3, 1)),

    dict(name="Architecture", tab="Architecture",
         owner="MVA", area="sol", level="Solution",
         target="Established",
         desc="De minimale set van structuren, bouwstenen en overzichten die nodig zijn om de significante drivers te vervullen en levering mogelijk te maken.",
         states=[
             ("Envisioned", [
                 "De doelrichting, scope en waarde zijn overeengekomen op één pagina.",
                 "De sleutelstakeholders delen de visie.",
             ]),
             ("Outlined", [
                 "De minimale haalbare structuur, grenzen en interfaces zijn gedefinieerd.",
                 "Alleen de bouwstenen die nu nodig zijn zijn opgenomen.",
             ]),
             ("Demonstrated", [
                 "De risicovolste delen zijn bewezen via spike, referentie of increment.",
                 "De sleutelbeslissingen zijn gevalideerd door bewijs.",
             ]),
             ("Usable", [
                 "De architectuur leidt actief levering en stelt beperkingen.",
                 "De teams bouwen er actief binnen.",
             ]),
             ("Established", [
                 "De architectuur is operationeel en levert de beoogde uitkomsten.",
             ]),
             ("Evolving", [
                 "De architectuur verandert onder beheerde, continue aanpassing.",
                 "Wijziging wordt afgehandeld zonder grootschalige herontwikkeling.",
             ]),
         ], sample=(3, 1)),

    dict(name="Architecture Decisions", tab="Architecture Decisions",
         owner="MVA", area="sol", level="Solution",
         target="Validated",
         desc="De significante, moeilijk terug te draaien keuzes die de architectuur vormen.",
         states=[
             ("Needed", ["Een significant, moeilijk terug te draaien beslissingspunt is herkend en eigendom aangewezen."]),
             ("Framed", ["De opties, afwegingen en het laatste verantwoorde moment zijn begrepen."]),
             ("Decided", ["De keuze is gemaakt en vastgelegd met onderbouwing en consequenties (ADR)."]),
             ("Communicated", ["De beslissing is bekend en geaccepteerd door degenen die het beïnvloedt."]),
             ("Validated", ["De beslissing is bewezen door bewijs en uitkomst, of herzien."]),
         ], sample=(4, 0)),

    dict(name="Paved Road", tab="Paved Road",
         owner="MVA", area="sol", level="Enterprise",
         target="Pervasive",
         desc="Het herbruikbare servicechassis van goedgekeurde producten, patronen en standaarden dat de conforme keuze de gemakkelijke maakt.",
         states=[
             ("Seeded", [
                 "De behoefte aan een paved road is erkend.",
                 "Een eerste goedgekeurde standaard of twee bestaat.",
             ]),
             ("Established", [
                 "Een coherent servicechassis van goedgekeurde producten, patronen en standaarden is gepubliceerd.",
                 "De paved road is zelfbedienbaar.",
             ]),
             ("Adopted", [
                 "Teams bouwen standaard op de paved road.",
                 "On-road keuzes hebben geen beoordeling nodig.",
             ]),
             ("Pervasive", [
                 "De paved road is de norm over het landgoed.",
                 "Off-road gaan is de zeldzame, gerechtvaardige uitzondering.",
             ]),
             ("Evolving", [
                 "De paved road wordt continu verbreed en gesnoeid.",
                 "Lessen uit beoordelingen worden teruggekoppeld.",
             ]),
         ], sample=(2, 1)),

    # ── ENDEAVOR ──────────────────────────────────────────────────────────
    dict(name="Work", tab="Work",
         owner="Kernel", area="end", level="Solution",
         target="Closed",
         desc="De activiteit en inspanning die worden ingezet om het resultaat te produceren.",
         states=[
             ("Initiated", [
                 "De beperkingen en financieringscontext zijn duidelijk.",
                 "De sponsors zijn geïdentificeerd.",
                 "De prioriteit is duidelijk.",
                 "Het werk is aangevraagd.",
             ]),
             ("Prepared", [
                 "Een geloofwaardig plan en schatting zijn aanwezig.",
                 "Financiering en middelen zijn toegezegd.",
                 "Sleutelrisico's zijn geïdentificeerd en worden aangepakt.",
                 "De succescriteria zijn overeengekomen.",
                 "Governance en de manier van werken zijn aanwezig.",
             ]),
             ("Started", [
                 "Het werk is begonnen en voortgang wordt bewaakt.",
                 "Het team voert het werk uit.",
                 "Taken worden afgerond.",
             ]),
             ("Under Control", [
                 "Taken worden afgerond en risico's zijn onder controle.",
                 "Schattingen worden bijgewerkt en herwerk is onder controle.",
                 "Meetgegevens van voortgang en snelheid zijn beschikbaar.",
             ]),
             ("Concluded", [
                 "Het werk om de resultaten te produceren is afgerond.",
                 "De resultaten zijn bereikt en geaccepteerd door de stakeholders.",
             ]),
             ("Closed", [
                 "Alles is gearchiveerd en lessen zijn vastgelegd.",
                 "Het budget is afgerekend en het team is vrijgegeven.",
             ]),
         ], sample=(2, 2)),

    dict(name="Team", tab="Team",
         owner="Kernel", area="end", level="Solution",
         target="Performing",
         desc="De groep die het systeem actief ontwikkelt, onderhoudt en ondersteunt.",
         states=[
             ("Seeded", [
                 "De teammissie is gedefinieerd.",
                 "De beperkingen voor het team zijn bekend.",
                 "De vereiste competenties zijn geïdentificeerd.",
                 "Het leiderschap- en governance-model is geselecteerd.",
                 "De teamomvang en -samenstelling zijn overeengekomen.",
             ]),
             ("Formed", [
                 "Voldoende teamleden zijn geworven om te beginnen.",
                 "Rollen en verantwoordelijkheden zijn begrepen.",
                 "De leden begrijpen hoe ze samen moeten werken.",
                 "De leden zijn betrokken en toegewijd.",
             ]),
             ("Collaborating", [
                 "Het team werkt als één samenhangend geheel.",
                 "Communicatie is open en eerlijk.",
                 "Het team is gericht op het bereiken van de missie.",
             ]),
             ("Performing", [
                 "Het team komt zijn verplichtingen consistent na.",
                 "Het team past zich aan de veranderende context aan.",
                 "Het team identificeert en pakt problemen aan zonder externe hulp.",
             ]),
             ("Adjourned", [
                 "De verantwoordelijkheden van het team zijn overgedragen of vervuld.",
                 "De middelen zijn vrijgegeven.",
             ]),
         ], sample=(2, 2)),

    dict(name="Way of Working", tab="Way of Working",
         owner="Kernel", area="end", level="Solution",
         target="Working Well",
         desc="De op maat gemaakte set van practices en tools die het team gebruikt om zijn werk te sturen.",
         states=[
             ("Principles Established", [
                 "De principes en beperkingen zijn door het team aanvaard.",
                 "De principes zijn overeengekomen met de stakeholders.",
                 "De tool- en practicebehoeften zijn begrepen.",
             ]),
             ("Foundation Established", [
                 "De sleutelpractices en tools zijn geselecteerd en voldoende geïntegreerd om te beginnen.",
                 "De niet-onderhandelbare practices zijn overeengekomen.",
                 "De capaciteitslacunes zijn begrepen.",
             ]),
             ("In Use", [
                 "De practices en tools worden gebruikt door het team.",
                 "Het team past de manier van werken aan aan zijn behoeften.",
                 "Feedbackmechanismen zijn aanwezig.",
             ]),
             ("In Place", [
                 "Het volledige team gebruikt de manier van werken.",
                 "Het team heeft toegang tot de tools.",
                 "De manier van werken wordt regelmatig geïnspecteerd.",
             ]),
             ("Working Well", [
                 "Het team levert voorspelbaar op met behulp van de manier van werken.",
                 "De practices worden natuurlijk toegepast en continu afgestemd.",
                 "Het team verbetert de manier van werken continu.",
             ]),
             ("Retired", [
                 "De manier van werken is niet meer in gebruik.",
                 "De geleerde lessen zijn vastgelegd.",
             ]),
         ], sample=(2, 1)),

    dict(name="Governance", tab="Governance",
         owner="MVG", area="end", level="Spans",
         target="Effective",
         desc="De lichtgewichte beslissingsbevoegdheden, guardrails en zekerheid die levering in lijn houden met de bedoeling.",
         states=[
             ("Initiated", [
                 "De beslissingsbevoegdheden en minimale guardrails zijn overeengekomen.",
                 "Het is duidelijk wie wat beslist.",
             ]),
             ("Engaged", [
                 "Een lean governance-ritme is operationeel.",
                 "De teams kennen de guardrails.",
             ]),
             ("Guiding", [
                 "Beslissingen stromen binnen de guardrails.",
                 "Alleen uitzonderingen worden geëscaleerd.",
             ]),
             ("Assured", [
                 "Conformance is aangetoond waar het werkelijk van belang is.",
                 "De architectuurbeoordelingen worden opgevolgd.",
             ]),
             ("Effective", [
                 "Governance maakt levering aantoonbaar mogelijk.",
                 "De governance zelf wordt continu afgestemd en verlicht.",
             ]),
         ], sample=(3, 1)),
]

# ─── Stijlhulpen ─────────────────────────────────────────────────────────────

def solid(hex6):
    """PatternFill met volledig dekkende kleur (ARGB prefix FF)."""
    argb = hex6 if len(hex6) == 8 else "FF" + hex6
    return PatternFill("solid", fgColor=argb)

def fnt(size=10, bold=False, color=INK, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def aln(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap,
                     indent=indent, shrink_to_fit=False)

def hair():
    return Side(style="hair", color="D8D0BF")

def thin_white():
    return Side(style="thin", color="FFFFFF")

def no_side():
    return Side(style=None)

def hair_box():
    h = hair()
    return Border(left=h, right=h, top=h, bottom=h)

def hair_bot():
    return Border(bottom=hair())

def set_w(ws, col, w, hidden=False):
    ws.column_dimensions[col].width = w
    if hidden:
        ws.column_dimensions[col].hidden = True


# ─── Dashboard ────────────────────────────────────────────────────────────────

def build_dashboard(wb, alphas):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK
    ws.freeze_panes = "A9"

    for col, w in [("A",3),("B",24),("C",15),("D",13),("E",22),
                   ("F",16),("G",22),("H",18)]:
        set_w(ws, col, w)
    set_w(ws, "O", 13, hidden=True)

    tab_names = [a["tab"] for a in alphas]

    def ref(tab):
        return f"'{tab}'" if " " in tab else tab

    # Rij 1: spacer
    ws.row_dimensions[1].height = 15

    # Rij 2: Titel
    ws.row_dimensions[2].height = 45.75
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "Essence Alpha Assessment"
    c.font = fnt(26, bold=True, color=WHITE)
    c.fill = solid(DARK)
    c.alignment = aln(indent=1)

    # Rij 3: Ondertitel
    ws.row_dimensions[3].height = 21.75
    ws.merge_cells("B3:H3")
    c = ws["B3"]
    c.value = ("Projectvoortgang & prestaties — kernel + practices, gemeten door alpha-toestand  ·  "
               "beoordeeld op solution-niveau (zie het tabblad Schaalniveaus)")
    c.font = fnt(11, color=GRAY)
    c.fill = solid(DARK)
    c.alignment = aln(wrap=True, indent=1)

    # Rij 4: spacer
    ws.row_dimensions[4].height = 15

    # Rij 5: KPI-labels (3 gekleurde tegels)
    ws.row_dimensions[5].height = 18
    kpi_colors = [MVA_C, MVG_C, KRN_C]
    kpi_labels = ["TOTALE VOLTOOIING", "ALPHAS VOLLEDIG", "CHECKLISTPUNTEN"]
    for col, lbl, clr in zip(["B","D","F"], kpi_labels, kpi_colors):
        c = ws[f"{col}5"]
        c.value = lbl
        c.font = fnt(9, bold=True, color=WHITE)
        c.fill = solid(clr)
        c.alignment = aln(indent=1)

    # Rij 6: KPI-waarden (worden later ingevuld als rijen bekend zijn)
    ws.row_dimensions[6].height = 39.75

    # Rij 7: spacer
    ws.row_dimensions[7].height = 7

    # Rij 8: Tabelkoptekst
    ws.row_dimensions[8].height = 19.5
    for col, lbl in [("B","ALPHA"),("C","EIGENAAR"),("D","NIVEAU"),
                     ("E","HUIDIGE TOESTAND"),("F","VOORTGANG"),
                     ("G","DOELTOESTAND"),("H","STATUS")]:
        c = ws[f"{col}8"]
        c.value = lbl
        c.font = fnt(9, bold=True, color=WHITE)
        c.fill = solid(DARK)
        c.alignment = aln(indent=1)

    # Alpha-rijen
    row = 9
    status_rows = []

    area_groups = [("CUSTOMER","cust",CUST), ("SOLUTION","sol",SOL),
                   ("ENDEAVOR","end",END)]

    for area_label, area_key, area_color in area_groups:
        ws.row_dimensions[row].height = 19.5
        ws.merge_cells(f"B{row}:H{row}")
        c = ws[f"B{row}"]
        c.value = area_label
        c.font = fnt(10, bold=True, color=WHITE)
        c.fill = solid(area_color)
        c.alignment = aln(indent=2)
        row += 1

        for a in [x for x in alphas if x["area"] == area_key]:
            ws.row_dimensions[row].height = 18
            t = a["tab"]
            r_ref = ref(t)

            # B: alpha naam (vet, eigenaar-kleur, hyperlink)
            b = ws.cell(row=row, column=2)
            b.value = a["name"]
            b.hyperlink = f"#{t}!B2"
            b.font = Font(name="Calibri", size=10.5, bold=True,
                          color=OWN_COL[a["owner"]])
            b.alignment = aln(indent=1)

            # C: eigenaar-chip
            ws.cell(row=row, column=3).value = a["owner"]
            ws.cell(row=row, column=3).font = fnt(8.5, bold=True, color=WHITE)
            ws.cell(row=row, column=3).fill = solid(OWN_COL[a["owner"]])
            ws.cell(row=row, column=3).alignment = aln(h="center")

            # D: niveau-chip
            lbl = LVL_LBL[a["level"]]
            ws.cell(row=row, column=4).value = lbl
            ws.cell(row=row, column=4).font = fnt(8.5, bold=True, color=WHITE)
            ws.cell(row=row, column=4).fill = solid(LVL_COL[a["level"]])
            ws.cell(row=row, column=4).alignment = aln(h="center")

            # E: huidige toestand
            ws.cell(row=row, column=5).value = f"={r_ref}!B5"
            ws.cell(row=row, column=5).font = fnt(10, color=INK)
            ws.cell(row=row, column=5).alignment = aln(indent=1)

            # F: voortgang %
            ws.cell(row=row, column=6).value = f"={r_ref}!C5"
            ws.cell(row=row, column=6).number_format = "0%"
            ws.cell(row=row, column=6).font = fnt(10, bold=True, color=GRAY)
            ws.cell(row=row, column=6).alignment = aln(h="center")

            # G: doeltoestand
            ws.cell(row=row, column=7).value = a["target"]
            ws.cell(row=row, column=7).font = fnt(9.5, italic=True, color=GRAY)
            ws.cell(row=row, column=7).alignment = aln(indent=1)

            # H: status-formule
            ws.cell(row=row, column=8).value = (
                f'=IF(F{row}>=1,"✔ Volledig",'
                f'IF(F{row}>0,"● In uitvoering","○ Niet gestart"))')
            ws.cell(row=row, column=8).font = fnt(9.5, bold=True)
            ws.cell(row=row, column=8).alignment = aln(indent=1)

            status_rows.append(row)
            row += 1

    # KPI-waarden (rij 6) nu rijen bekend zijn
    s0, sn = status_rows[0], status_rows[-1]
    n = len(alphas)

    kpi_vals = [
        ("B", "=O1",            "0%",  MVA_C),
        ("D", f'=COUNTIF(H{s0}:H{sn},"✔ Volledig")&" / "&O3', "@", MVG_C),
        ("F", "=O4&\" / \"&O5", "@",  KRN_C),
    ]
    for col, formula, fmt, clr in kpi_vals:
        c = ws[f"{col}6"]
        c.value = formula
        c.number_format = fmt
        c.font = fnt(20, bold=True, color=clr)
        c.fill = solid(tint(clr))
        c.alignment = aln(indent=1, v="center")
    ws.row_dimensions[6].height = 39.75

    # Verborgen O-kolom (aggregaten)
    m1s = "+".join(f"{ref(t)}!M1" for t in tab_names)
    m2s = "+".join(f"{ref(t)}!M2" for t in tab_names)
    ws["O1"].value = "=O4/O5"
    ws["O3"].value = n
    ws["O4"].value = f"={m1s}"
    ws["O5"].value = f"={m2s}"

    # Databalk op voortgang-kolom F
    ws.conditional_formatting.add(
        f"F{s0}:F{sn}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1, color=MVA_C))

    # Voettekst
    foot_row = row + 1
    ws.row_dimensions[foot_row].height = 30
    ws.merge_cells(f"B{foot_row}:H{foot_row}")
    c = ws[f"B{foot_row}"]
    c.value = ("Tip: open een alpha-tabblad en vink de ✔-kolom (dropdown) aan per checklistpunt. "
               "Huidige toestand, voortgang en dit dashboard worden automatisch bijgewerkt.  •  "
               "Voorbeelddata weergegeven — wis de ✔-cellen voor een lege beoordeling.")
    c.font = fnt(9, italic=True, color=GRAY)
    c.alignment = aln(wrap=True, indent=1)

    leg_row = foot_row + 1
    ws.row_dimensions[leg_row].height = 18
    ws.merge_cells(f"B{leg_row}:H{leg_row}")
    c = ws[f"B{leg_row}"]
    c.value = ("Status:  ✔ Volledig   ● In uitvoering   ○ Niet gestart        "
               "Niveau:  Enterprise (geërfd)  ·  Solution (eigendom)  ·  Spans")
    c.font = fnt(9, color=GRAY)
    c.alignment = aln(indent=1)

    ref_row = leg_row + 1
    ws.row_dimensions[ref_row].height = 16
    c = ws[f"B{ref_row}"]
    c.value = "→ Zie het tabblad 'Schaalniveaus' voor uitleg over de niveaus"
    c.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
    c.hyperlink = "#Schaalniveaus!B1"
    c.alignment = aln(indent=1)


# ─── Schaalniveaus-tabblad ────────────────────────────────────────────────────

def build_levels_tab(wb):
    ws = wb.create_sheet("Schaalniveaus")
    ws.sheet_properties.tabColor = KRN_C

    set_w(ws, "A", 3)
    set_w(ws, "B", 52)
    set_w(ws, "C", 52)
    set_w(ws, "D", 52)

    ws.row_dimensions[1].height = 28
    ws.merge_cells("B1:D1")
    c = ws["B1"]
    c.value = "Schaalniveaus — het System benoemen"
    c.font = fnt(14, bold=True, color=WHITE)
    c.fill = solid(KRN_C)
    c.alignment = aln(indent=1)

    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = ("Het System-alpha is het systeem van interesse — en het is fractaal. "
               "Kies het niveau voordat u beoordeelt.")
    c.font = fnt(10, color=WHITE)
    c.fill = solid(KRN_C)
    c.alignment = aln(wrap=True, indent=1)

    ws.row_dimensions[3].height = 54
    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = ("Een enterprise is een systeem van systemen; een solution is een sub-systeem van de enterprise; "
               "een geconfigureerd product is een sub-systeem van de solution. Dezelfde kernel en practices "
               "worden op elk niveau toegepast — en wat op één niveau een System is, is op het niveau daarboven "
               "een sub-systeem. De grens die u trekt bepaalt wat elk alpha werkelijk meet.")
    c.font = fnt(10, color="BBBBBB")
    c.fill = solid(DARK)
    c.alignment = aln(wrap=True, indent=1)

    levels = [
        ("B4", KRN_C, "ENTERPRISE — System = de enterprise\nEnterprise-architectuur, de paved road, drivers en staande governance zijn hier eigendom."),
        ("C4", MVA_C, "SOLUTION / PROJECT — System = de solution\nErft de enterprise paved road, drivers en governance als context; bezit de architectuur en beslissingen voor deze solution."),
        ("D4", "6B6256", "COMPONENT — System = een geconfigureerd product\nEen sub-systeem van de solution, apart bijgehouden wanneer dat gerechtvaardigd is."),
    ]
    ws.row_dimensions[4].height = 72
    for cell_ref, color, text in levels:
        c = ws[cell_ref]
        c.value = text
        c.font = fnt(10, color=WHITE)
        c.fill = solid(color)
        c.alignment = aln(v="top", wrap=True, indent=1)

    ws.row_dimensions[5].height = 18
    ws["B5"].value = "Dezelfde kernel + practices lopen op elk niveau; een System op één niveau is een sub-systeem op het niveau daarboven."
    ws["B5"].font = fnt(9, italic=True, color=GRAY)
    ws["B5"].alignment = aln(indent=1)

    ws.row_dimensions[6].height = 22
    ws["B6"].value = "WELKE PRACTICE WAAR LOOPT"
    ws["B6"].font = fnt(11, bold=True)
    ws["B6"].alignment = aln(indent=1)

    ws.row_dimensions[7].height = 18
    ws["B7"].value = "ENTERPRISE  →"
    ws["B7"].font = fnt(10, bold=True)
    ws["B7"].alignment = aln(indent=1)
    ws["C7"].value = "TOGAF Business Scenarios-techniek (kwantificeer de drivers)  ·  enterprise MVA  ·  MVG (guardrails + paved road) — doorgegeven als context"
    ws["C7"].font = fnt(10)
    ws["C7"].alignment = aln(wrap=True)

    ws.row_dimensions[8].height = 18
    ws["B8"].value = "SOLUTION  →"
    ws["B8"].font = fnt(10, bold=True)
    ws["B8"].alignment = aln(indent=1)
    ws["C8"].value = "MVA (architectureer binnen de guardrails) — stuurt bewijs en nieuwe drivers terug omhoog"
    ws["C8"].font = fnt(10)
    ws["C8"].alignment = aln(wrap=True)

    ws.row_dimensions[9].height = 40
    ws.merge_cells("B9:D9")
    c = ws["B9"]
    c.value = ("Dit werkboek beoordeelt op SOLUTION-niveau. Alphas die hier eigendom zijn zijn getagd als "
               "Solution; de Paved Road is Enterprise (geërfd van de Enterprise); Governance overspant niveaus. "
               "De niveau-tag staat op elk alpha-tabblad (rechtsbovenin) en in het dashboard.")
    c.font = fnt(9, italic=True, color=GRAY)
    c.alignment = aln(wrap=True, indent=1)


# ─── Alpha-tabblad ────────────────────────────────────────────────────────────

def build_alpha_tab(wb, a):
    tab    = a["tab"]
    owner  = a["owner"]
    level  = a["level"]
    states = a["states"]
    s_full, s_part = a["sample"]

    own_c  = OWN_COL[owner]
    lvl_c  = LVL_COL[level]
    kpi_bg = tint(own_c)   # lichte tint voor KPI-rij

    ws = wb.create_sheet(title=tab)
    ws.sheet_properties.tabColor = AREA_COL[a["area"]]

    # Kolombreedte — identiek origineel
    set_w(ws, "A", 3)
    set_w(ws, "B", 27)
    set_w(ws, "C", 6)
    set_w(ws, "D", 95)
    set_w(ws, "E", 16)
    set_w(ws, "G", 13, hidden=True)
    for col in list("HIJK"):
        set_w(ws, col, 3, hidden=True)
    set_w(ws, "M", 3, hidden=True)

    n_states = len(states)

    # ── Rij 1: Terug naar dashboard ──────────────────────────────────────────
    ws.row_dimensions[1].height = 15
    c = ws["B1"]
    c.value = "← Dashboard"
    c.hyperlink = "#Dashboard!A1"
    c.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
    c.alignment = aln(indent=1)

    # ── Rij 2: Titel + niveau-chip ───────────────────────────────────────────
    ws.row_dimensions[2].height = 39.75
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = a["name"]
    c.font = fnt(22, bold=True, color=WHITE)
    c.fill = solid(own_c)
    c.alignment = aln(v="center", indent=1)

    ws["E2"].value = LVL_LBL[level]
    ws["E2"].font = fnt(10, bold=True, color=WHITE)
    ws["E2"].fill = solid(lvl_c)
    ws["E2"].alignment = aln(h="center")

    # ── Rij 3: Beschrijving ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 31.5
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = a["desc"]
    c.font = fnt(10.5, color=GRAY)
    c.alignment = aln(wrap=True, v="center", indent=1)

    # ── Rij 4: KPI-labels ────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 13.5
    for col, lbl in [("B","HUIDIGE TOESTAND"),("C","VOORTGANG"),
                      ("D","TOESTANDEN BEREIKT"),("E","EIGENAAR · GEBIED")]:
        c = ws[f"{col}4"]
        c.value = lbl
        c.font = fnt(8, bold=True, color=GRAY)
        c.alignment = aln(v="bottom", indent=1)

    # ── Rij 5: KPI-waarden (formules worden later ingevuld) ──────────────────
    ws.row_dimensions[5].height = 25.5
    for col in ["B","C","D","E"]:
        ws[f"{col}5"].fill = solid(kpi_bg)
    ws["E5"].value = f"{owner} · {a['area'].upper()}"
    ws["E5"].font = fnt(9.5, bold=True, color=own_c)
    ws["E5"].alignment = aln(indent=1)

    # ── Rij 6: Tabelkoptekst ─────────────────────────────────────────────────
    ws.row_dimensions[6].height = 19.5
    for col, lbl in [("B","TOESTAND"),("C","✔"),("D","CHECKLISTPUNT"),
                      ("E","TOESTANDSSTATUS")]:
        c = ws[f"{col}6"]
        c.value = lbl
        c.font = fnt(9, bold=True, color=WHITE)
        c.fill = solid(DARK)
        c.alignment = aln(h="center" if col == "C" else "left",
                           indent=0 if col == "C" else 1)

    # ── Data-validatie ───────────────────────────────────────────────────────
    dv = DataValidation(type="list", formula1='"✔"', allow_blank=True,
                        showErrorMessage=False, showInputMessage=False)
    ws.add_data_validation(dv)

    # ── Rijen 7+: toestanden + helper-kolommen ───────────────────────────────
    data_row   = 7   # rijcursor in B-E data
    helper_row = 7   # rijcursor in G-K helpers (één per toestand)
    state_info = []  # (staatsnaam, c_start, c_end, helper_row)

    for s_idx, (state_name, items) in enumerate(states):
        n_items = len(items)
        c_start = data_row
        c_end   = data_row + n_items - 1
        state_info.append((state_name, c_start, c_end, helper_row))

        # ── Col B: staatsnaam (gemerged, eigenaar-kleur) ─────────────────────
        if n_items > 1:
            ws.merge_cells(f"B{data_row}:B{c_end}")
        b = ws.cell(row=data_row, column=2)
        b.value = f"{s_idx + 1}. {state_name}"
        b.font = fnt(11, bold=True, color=WHITE)
        b.fill = solid(own_c)
        b.alignment = aln(v="top", wrap=True, indent=1)

        # Borders op col B: witte scheidingslijn boven eerste rij van elke toestand
        # Alle rijen: dun wit onder. Eerste rij: ook dun wit boven.
        for i in range(n_items):
            r = data_row + i
            top_s = thin_white() if i == 0 else no_side()
            ws.cell(row=r, column=2).border = Border(
                top=top_s, bottom=thin_white())

        # ── Col E: toestandsstatus (gemerged, amber) ─────────────────────────
        if n_items > 1:
            ws.merge_cells(f"E{data_row}:E{c_end}")
        e = ws.cell(row=data_row, column=5)
        e.value = f'=COUNTIF(C{c_start}:C{c_end},"✔")&" / {n_items}"'
        e.font = fnt(11, bold=True, color=AMB_FG)
        e.fill = solid(AMB_BG)
        e.alignment = aln(h="center", wrap=True)

        # Conditionele opmaak: groen zodra alle punten van deze toestand aangevinkt
        ws.conditional_formatting.add(
            f"E{c_start}:E{c_end}",
            FormulaRule(
                formula=[f'COUNTIF($C${c_start}:$C${c_end},"✔")={n_items}'],
                font=Font(bold=True, color=CHK_GR)))

        # ── Col C + D: checklistpunten ───────────────────────────────────────
        for i, item_text in enumerate(items):
            r = data_row + i
            ws.row_dimensions[r].height = 21

            # Alternerend: oneven absolute rij = wit, even = warm-grijs (net als origineel)
            row_bg = WHITE if r % 2 == 1 else PAPER2

            c_cell = ws.cell(row=r, column=3)
            c_cell.value = ""
            c_cell.font = fnt(11, bold=True, color=CHK_GR)
            c_cell.fill = solid(row_bg)
            c_cell.alignment = aln(h="center")
            c_cell.border = hair_box()
            dv.add(c_cell)

            d_cell = ws.cell(row=r, column=4)
            d_cell.value = item_text
            d_cell.font = fnt(10, color=DARK)
            d_cell.fill = solid(row_bg)
            d_cell.alignment = aln(wrap=True, v="center", indent=1)
            d_cell.border = hair_bot()

        # ── Helper G-K (één rij per toestand, op helper_row) ─────────────────
        ws.cell(row=helper_row, column=7).value  = state_name
        ws.cell(row=helper_row, column=8).value  = (
            f'=COUNTIF(C{c_start}:C{c_end},"✔")')
        ws.cell(row=helper_row, column=9).value  = n_items
        ws.cell(row=helper_row, column=10).value = (
            f'=IF(H{helper_row}=I{helper_row},1,0)')
        if s_idx == 0:
            ws.cell(row=helper_row, column=11).value = f'=J{helper_row}'
        else:
            prev_k = state_info[s_idx - 1][3]
            ws.cell(row=helper_row, column=11).value = (
                f'=J{helper_row}*K{prev_k}')

        data_row   += n_items
        helper_row += 1

    last_h = 7 + n_states - 1

    # ── M1–M4 ────────────────────────────────────────────────────────────────
    ws["M1"].value = f"=SUM(H7:H{last_h})"
    ws["M2"].value = f"=SUM(I7:I{last_h})"
    ws["M3"].value = f"=SUM(K7:K{last_h})"
    ws["M4"].value = f"=SUM(J7:J{last_h})"

    # ── Rij 5: KPI-formules ──────────────────────────────────────────────────
    ws["B5"].value = f'=IF(M3=0,"Niet gestart",INDEX(G7:G{last_h},M3))'
    ws["B5"].font  = fnt(12, bold=True, color=own_c)
    ws["B5"].alignment = aln(indent=1)

    ws["C5"].value = "=IF(M2=0,0,M1/M2)"
    ws["C5"].number_format = "0%"
    ws["C5"].font  = fnt(14, bold=True, color=own_c)
    ws["C5"].alignment = aln(indent=1)

    ws["D5"].value = f'=M4&" / "&{n_states}'
    ws["D5"].font  = fnt(12, bold=True, color=own_c)
    ws["D5"].alignment = aln(indent=1)

    # ── Voorbeelddata ────────────────────────────────────────────────────────
    for s_idx, (_, c_start, c_end, _h) in enumerate(state_info):
        if s_idx < s_full:
            for r in range(c_start, c_end + 1):
                ws.cell(row=r, column=3).value = "✔"
        elif s_idx == s_full and s_part > 0:
            for r in range(c_start, min(c_start + s_part, c_end + 1)):
                ws.cell(row=r, column=3).value = "✔"
            break


# ─── Hoofdprogramma ──────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    build_dashboard(wb, ALPHAS)
    build_levels_tab(wb)
    for a in ALPHAS:
        build_alpha_tab(wb, a)

    out = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "docs",
                     "EA_Alpha_Assessment.xlsx"))
    wb.save(out)

    total = sum(sum(len(items) for _, items in a["states"]) for a in ALPHAS)
    print(f"Opgeslagen: {out}")
    print(f"Alphas: {len(ALPHAS)}  |  Tabbladen: {len(wb.worksheets)}  "
          f"|  Checklistpunten: {total}")


if __name__ == "__main__":
    main()
